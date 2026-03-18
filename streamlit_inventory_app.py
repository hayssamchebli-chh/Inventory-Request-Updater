import io
from pathlib import Path
from typing import Dict, Optional, Tuple

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Inventory Request Updater", page_icon="📦", layout="centered")

YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF59D")
RED_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")


def normalize(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_number(value) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def find_header_row(ws, required_headers: list[str]) -> int:
    required = {h.casefold() for h in required_headers}

    for row_idx in range(1, min(ws.max_row, 50) + 1):
        row_values = {
            normalize(cell.value).casefold()
            for cell in ws[row_idx]
            if normalize(cell.value)
        }
        if required.issubset(row_values):
            return row_idx

    raise ValueError(
        f"Could not find a header row containing all required headers: {required_headers}"
    )


def build_column_map(ws, header_row: int) -> Dict[str, int]:
    return {
        normalize(cell.value): cell.column
        for cell in ws[header_row]
        if normalize(cell.value)
    }


def find_or_create_lead_time_column(ws, header_row: int) -> int:
    for cell in ws[header_row]:
        if normalize(cell.value).casefold() == "lead time delivery":
            return cell.column

    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col, value="Lead Time Delivery")
    return new_col


def clear_row_fills(ws, row_idx: int, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type=None)


def fill_row(ws, row_idx: int, max_col: int, fill: PatternFill) -> None:
    for col_idx in range(1, max_col + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def process_files(request_file, stock_file) -> bytes:
    request_wb = load_workbook(request_file)
    stock_wb = load_workbook(stock_file, data_only=True)

    request_ws = request_wb[request_wb.sheetnames[0]]
    stock_ws = stock_wb[stock_wb.sheetnames[0]]

    request_header_row = find_header_row(request_ws, ["No.2", "Qty Requested"])
    stock_header_row = find_header_row(
        stock_ws,
        ["Item No.1", "Stock Available Quantity", "Lead Time Delivery"],
    )

    request_cols = build_column_map(request_ws, request_header_row)
    stock_cols = build_column_map(stock_ws, stock_header_row)

    req_item_col = request_cols["No.2"]
    req_qty_col = request_cols["Qty Requested"]
    lead_time_col = find_or_create_lead_time_column(request_ws, request_header_row)

    stock_item_col = stock_cols["Item No.1"]
    stock_qty_col = stock_cols["Stock Available Quantity"]
    stock_lead_time_col = stock_cols["Lead Time Delivery"]

    stock_data: Dict[str, Tuple[Optional[float], object]] = {}
    for row_idx in range(stock_header_row + 1, stock_ws.max_row + 1):
        item_no = normalize(stock_ws.cell(row=row_idx, column=stock_item_col).value)
        if not item_no:
            continue
        available_qty = to_number(stock_ws.cell(row=row_idx, column=stock_qty_col).value)
        lead_time = stock_ws.cell(row=row_idx, column=stock_lead_time_col).value
        stock_data[item_no] = (available_qty, lead_time)

    max_used_col = max(request_ws.max_column, lead_time_col)

    for row_idx in range(request_header_row + 1, request_ws.max_row + 1):
        item_no = normalize(request_ws.cell(row=row_idx, column=req_item_col).value)
        requested_qty = to_number(request_ws.cell(row=row_idx, column=req_qty_col).value)

        clear_row_fills(request_ws, row_idx, max_used_col)
        request_ws.cell(row=row_idx, column=lead_time_col).value = None

        if not item_no:
            continue

        if item_no not in stock_data:
            fill_row(request_ws, row_idx, max_used_col, RED_FILL)
            continue

        available_qty, lead_time = stock_data[item_no]
        request_ws.cell(row=row_idx, column=lead_time_col, value=lead_time)

        if (
            requested_qty is not None
            and available_qty is not None
            and requested_qty < available_qty
        ):
            fill_row(request_ws, row_idx, max_used_col, YELLOW_FILL)

    request_ws.column_dimensions[
        request_ws.cell(row=request_header_row, column=lead_time_col).column_letter
    ].width = 20

    output = io.BytesIO()
    request_wb.save(output)
    output.seek(0)
    return output.getvalue()


st.title("📦 Inventory Request Updater")
st.write(
    "Upload the request file and the stock file. The app will update the request file, "
    "add **Lead Time Delivery**, highlight matching rows in yellow when stock is sufficient, "
    "and mark missing items in red."
)

request_file = st.file_uploader(
    "Upload the request quantity Excel file",
    type=["xlsx"],
    key="request_file",
)

stock_file = st.file_uploader(
    "Upload the available stock Excel file",
    type=["xlsx"],
    key="stock_file",
)

if request_file and stock_file:
    if st.button("Process files"):
        try:
            updated_file = process_files(request_file, stock_file)
            output_name = f"updated_{Path(request_file.name).stem}.xlsx"

            st.success("File processed successfully.")
            st.download_button(
                label="Download updated Excel file",
                data=updated_file,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Error: {e}")
else:
    st.info("Please upload both Excel files to continue.")
